import streamlit as st
import requests
import xml.etree.ElementTree as ET
import pandas as pd
import io
from datetime import date, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Tally Data Extractor",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
#  CUSTOM CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
    /* Main background */
    .stApp { background-color: #0f1117; }

    /* Sidebar */
    section[data-testid="stSidebar"] {
        background: #161b27;
        border-right: 1px solid #252d3d;
    }

    /* Cards */
    .tally-card {
        background: #161b27;
        border: 1px solid #252d3d;
        border-radius: 12px;
        padding: 20px 24px;
        margin-bottom: 16px;
    }
    .tally-card-title {
        font-size: 11px;
        font-weight: 700;
        letter-spacing: 2px;
        text-transform: uppercase;
        color: #6b7280;
        margin-bottom: 14px;
    }

    /* Status badges */
    .badge-connected {
        display: inline-block;
        background: rgba(79,255,176,.15);
        color: #4fffb0;
        border: 1px solid rgba(79,255,176,.3);
        border-radius: 20px;
        padding: 3px 12px;
        font-size: 12px;
        font-weight: 600;
    }
    .badge-disconnected {
        display: inline-block;
        background: rgba(255,107,107,.15);
        color: #ff6b6b;
        border: 1px solid rgba(255,107,107,.3);
        border-radius: 20px;
        padding: 3px 12px;
        font-size: 12px;
        font-weight: 600;
    }

    /* Metric boxes */
    .metric-box {
        background: #1c2236;
        border: 1px solid #252d3d;
        border-radius: 10px;
        padding: 16px 18px;
        text-align: center;
    }
    .metric-value {
        font-size: 28px;
        font-weight: 800;
        color: #4fffb0;
        line-height: 1;
    }
    .metric-label {
        font-size: 11px;
        color: #6b7280;
        margin-top: 4px;
        text-transform: uppercase;
        letter-spacing: 1px;
    }

    /* Table styling */
    .stDataFrame { border-radius: 10px; overflow: hidden; }

    /* Buttons */
    .stButton > button {
        border-radius: 8px;
        font-weight: 600;
        transition: all .2s;
    }

    /* Info box */
    .info-box {
        background: rgba(91,143,255,.1);
        border: 1px solid rgba(91,143,255,.3);
        border-radius: 8px;
        padding: 12px 16px;
        font-size: 13px;
        color: #93b4ff;
    }

    /* Warning box */
    .warn-box {
        background: rgba(255,200,80,.1);
        border: 1px solid rgba(255,200,80,.3);
        border-radius: 8px;
        padding: 12px 16px;
        font-size: 13px;
        color: #ffd166;
    }

    h1, h2, h3 { color: #e8eaf0 !important; }
    p, label, .stMarkdown { color: #c0c4d0; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
#  TALLY XML HELPERS
# ─────────────────────────────────────────────
TALLY_URL_TEMPLATE = "http://{host}:{port}"

def tally_post(xml_body: str, host: str, port: int) -> str | None:
    """POST an XML request to Tally and return raw response text."""
    url = TALLY_URL_TEMPLATE.format(host=host, port=port)
    try:
        r = requests.post(url, data=xml_body.encode("utf-8"),
                          headers={"Content-Type": "application/xml"},
                          timeout=10)
        r.raise_for_status()
        return r.text
    except requests.exceptions.ConnectionError:
        return None
    except Exception as e:
        st.error(f"Tally request error: {e}")
        return None


def build_company_list_xml() -> str:
    return """<ENVELOPE>
  <HEADER>
    <TALLYREQUEST>Export Data</TALLYREQUEST>
  </HEADER>
  <BODY>
    <EXPORTDATA>
      <REQUESTDESC>
        <REPORTNAME>List of Companies</REPORTNAME>
        <STATICVARIABLES>
          <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
        </STATICVARIABLES>
      </REQUESTDESC>
    </EXPORTDATA>
  </BODY>
</ENVELOPE>"""


def build_ledger_xml(company: str, from_date: str, to_date: str) -> str:
    return f"""<ENVELOPE>
  <HEADER>
    <TALLYREQUEST>Export Data</TALLYREQUEST>
  </HEADER>
  <BODY>
    <EXPORTDATA>
      <REQUESTDESC>
        <REPORTNAME>Ledger</REPORTNAME>
        <STATICVARIABLES>
          <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
          <SVCURRENTCOMPANY>{company}</SVCURRENTCOMPANY>
          <SVFROMDATE>{from_date}</SVFROMDATE>
          <SVTODATE>{to_date}</SVTODATE>
        </STATICVARIABLES>
      </REQUESTDESC>
    </EXPORTDATA>
  </BODY>
</ENVELOPE>"""


def build_pl_xml(company: str, from_date: str, to_date: str) -> str:
    return f"""<ENVELOPE>
  <HEADER>
    <TALLYREQUEST>Export Data</TALLYREQUEST>
  </HEADER>
  <BODY>
    <EXPORTDATA>
      <REQUESTDESC>
        <REPORTNAME>Profit and Loss</REPORTNAME>
        <STATICVARIABLES>
          <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
          <SVCURRENTCOMPANY>{company}</SVCURRENTCOMPANY>
          <SVFROMDATE>{from_date}</SVFROMDATE>
          <SVTODATE>{to_date}</SVTODATE>
        </STATICVARIABLES>
      </REQUESTDESC>
    </EXPORTDATA>
  </BODY>
</ENVELOPE>"""


def build_bs_xml(company: str, as_on_date: str) -> str:
    return f"""<ENVELOPE>
  <HEADER>
    <TALLYREQUEST>Export Data</TALLYREQUEST>
  </HEADER>
  <BODY>
    <EXPORTDATA>
      <REQUESTDESC>
        <REPORTNAME>Balance Sheet</REPORTNAME>
        <STATICVARIABLES>
          <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
          <SVCURRENTCOMPANY>{company}</SVCURRENTCOMPANY>
          <SVTODATE>{as_on_date}</SVTODATE>
        </STATICVARIABLES>
      </REQUESTDESC>
    </EXPORTDATA>
  </BODY>
</ENVELOPE>"""


def build_voucher_xml(company: str, from_date: str, to_date: str) -> str:
    return f"""<ENVELOPE>
  <HEADER>
    <TALLYREQUEST>Export Data</TALLYREQUEST>
  </HEADER>
  <BODY>
    <EXPORTDATA>
      <REQUESTDESC>
        <REPORTNAME>Voucher Register</REPORTNAME>
        <STATICVARIABLES>
          <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
          <SVCURRENTCOMPANY>{company}</SVCURRENTCOMPANY>
          <SVFROMDATE>{from_date}</SVFROMDATE>
          <SVTODATE>{to_date}</SVTODATE>
        </STATICVARIABLES>
      </REQUESTDESC>
    </EXPORTDATA>
  </BODY>
</ENVELOPE>"""


# ─────────────────────────────────────────────
#  XML PARSERS  →  DataFrame
# ─────────────────────────────────────────────
def parse_companies(xml_text: str) -> list[str]:
    """Extract open company names from Tally XML response."""
    companies = []
    try:
        root = ET.fromstring(xml_text)
        for elem in root.iter():
            if elem.tag.upper() in ("COMPANY", "COMPANYNAME", "NAME"):
                if elem.text and elem.text.strip():
                    companies.append(elem.text.strip())
        # deduplicate while preserving order
        seen = set()
        unique = []
        for c in companies:
            if c not in seen:
                seen.add(c)
                unique.append(c)
        return unique
    except ET.ParseError:
        return []


def _safe_text(elem) -> str:
    return (elem.text or "").strip() if elem is not None else ""


def parse_ledgers(xml_text: str) -> pd.DataFrame:
    """Parse ledger XML into a flat DataFrame."""
    rows = []
    try:
        root = ET.fromstring(xml_text)
        for ledger in root.iter("LEDGER"):
            name   = _safe_text(ledger.find("NAME"))
            group  = _safe_text(ledger.find("PARENT"))
            opbal  = _safe_text(ledger.find("OPENINGBALANCE"))
            clbal  = _safe_text(ledger.find("CLOSINGBALANCE"))
            gstin  = _safe_text(ledger.find("GSTIN"))
            addr   = _safe_text(ledger.find("ADDRESS"))
            pan    = _safe_text(ledger.find("INCOMETAXNUMBER"))
            email  = _safe_text(ledger.find("EMAIL"))
            phone  = _safe_text(ledger.find("LEDPHONE"))
            rows.append({
                "Ledger Name": name,
                "Group / Parent": group,
                "Opening Balance": opbal,
                "Closing Balance": clbal,
                "GSTIN": gstin,
                "PAN": pan,
                "Email": email,
                "Phone": phone,
                "Address": addr,
            })
    except ET.ParseError as e:
        st.error(f"XML parse error (Ledger): {e}")
    return pd.DataFrame(rows)


def parse_pl(xml_text: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Parse P&L XML → (income_df, expense_df)."""
    income_rows, expense_rows = [], []
    try:
        root = ET.fromstring(xml_text)
        for group in root.iter("GROUP"):
            gname  = _safe_text(group.find("NAME"))
            amount = _safe_text(group.find("CLOSINGBALANCE"))
            nature = _safe_text(group.find("NATURE"))
            for ledger in group.iter("LEDGER"):
                lname  = _safe_text(ledger.find("NAME"))
                lamount = _safe_text(ledger.find("CLOSINGBALANCE"))
                row = {"Group": gname, "Ledger": lname, "Amount": lamount, "Nature": nature}
                if "income" in nature.lower() or "revenue" in nature.lower():
                    income_rows.append(row)
                else:
                    expense_rows.append(row)
    except ET.ParseError as e:
        st.error(f"XML parse error (P&L): {e}")
    return pd.DataFrame(income_rows), pd.DataFrame(expense_rows)


def parse_bs(xml_text: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Parse Balance Sheet XML → (liabilities_df, assets_df)."""
    liab_rows, asset_rows = [], []
    try:
        root = ET.fromstring(xml_text)
        for group in root.iter("GROUP"):
            gname  = _safe_text(group.find("NAME"))
            nature = _safe_text(group.find("NATURE"))
            for ledger in group.iter("LEDGER"):
                lname   = _safe_text(ledger.find("NAME"))
                lamount = _safe_text(ledger.find("CLOSINGBALANCE"))
                row = {"Group": gname, "Ledger": lname, "Closing Balance": lamount}
                if "asset" in nature.lower():
                    asset_rows.append(row)
                else:
                    liab_rows.append(row)
    except ET.ParseError as e:
        st.error(f"XML parse error (Balance Sheet): {e}")
    return pd.DataFrame(liab_rows), pd.DataFrame(asset_rows)


def parse_vouchers(xml_text: str) -> pd.DataFrame:
    """Parse Voucher Register XML → DataFrame."""
    rows = []
    try:
        root = ET.fromstring(xml_text)
        for v in root.iter("VOUCHER"):
            rows.append({
                "Date":          _safe_text(v.find("DATE")),
                "Voucher No":    _safe_text(v.find("VOUCHERNUMBER")),
                "Voucher Type":  _safe_text(v.find("VOUCHERTYPENAME")),
                "Party Name":    _safe_text(v.find("PARTYLEDGERNAME")),
                "Amount":        _safe_text(v.find("AMOUNT")),
                "Narration":     _safe_text(v.find("NARRATION")),
                "Ref No":        _safe_text(v.find("REFERENCE")),
            })
    except ET.ParseError as e:
        st.error(f"XML parse error (Vouchers): {e}")
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────
#  EXCEL EXPORT WITH FORMATTING
# ─────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1A2340")
HEADER_FONT   = Font(name="Arial", bold=True, color="4FFFB0", size=11)
ALT_FILL      = PatternFill("solid", fgColor="161B27")
NORMAL_FILL   = PatternFill("solid", fgColor="0F1117")
CELL_FONT     = Font(name="Arial", size=10, color="C0C4D0")
TITLE_FONT    = Font(name="Arial", bold=True, size=14, color="FFFFFF")
THIN_BORDER   = Border(
    bottom=Side(style="thin", color="252D3D"),
    right=Side(style="thin", color="252D3D"),
)

def _write_sheet(ws, title: str, df: pd.DataFrame, columns: list[str] | None = None):
    """Write a DataFrame to a worksheet with professional formatting."""
    if columns:
        df = df[[c for c in columns if c in df.columns]]

    # Title row
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(df.columns), 1))
    title_cell = ws.cell(1, 1)
    title_cell.font = TITLE_FONT
    title_cell.fill = PatternFill("solid", fgColor="0A3D2E")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Generated on row
    ws.append([f"Generated on: {datetime.now().strftime('%d-%b-%Y  %H:%M:%S')}"])
    ws.cell(2, 1).font = Font(name="Arial", size=9, color="6B7280", italic=True)
    ws.append([])  # blank spacer

    if df.empty:
        ws.append(["No data returned from Tally for this request."])
        ws.cell(ws.max_row, 1).font = Font(name="Arial", size=10, color="FF6B6B", italic=True)
        return

    # Headers
    ws.append(list(df.columns))
    header_row = ws.max_row
    for col_idx, _ in enumerate(df.columns, 1):
        cell = ws.cell(header_row, col_idx)
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 22

    # Data rows
    for r_idx, row_data in enumerate(df.itertuples(index=False), 1):
        ws.append(list(row_data))
        data_row = ws.max_row
        fill = ALT_FILL if r_idx % 2 == 0 else NORMAL_FILL
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(data_row, col_idx)
            cell.font   = CELL_FONT
            cell.fill   = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    # Auto-width columns
    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(col_name)), df[col_name].astype(str).map(len).max() if not df.empty else 0)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def build_excel(sheets: dict[str, tuple[str, pd.DataFrame, list | None]]) -> bytes:
    """
    sheets = { sheet_name: (display_title, dataframe, [column_order_or_None]) }
    Returns bytes of the Excel workbook.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for sheet_name, (title, df, columns) in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        _write_sheet(ws, title, df, columns)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────
#  SESSION STATE DEFAULTS
# ─────────────────────────────────────────────
if "connected"  not in st.session_state: st.session_state.connected  = False
if "companies"  not in st.session_state: st.session_state.companies  = []
if "host"       not in st.session_state: st.session_state.host       = "localhost"
if "port"       not in st.session_state: st.session_state.port       = 9000


# ─────────────────────────────────────────────
#  SIDEBAR  — CONNECTION + SETTINGS
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Tally Connection")
    st.divider()

    host = st.text_input("Tally Host / IP", value=st.session_state.host,
                         placeholder="localhost")
    port = st.number_input("Tally Port", value=st.session_state.port,
                           min_value=1, max_value=65535, step=1)

    col_a, col_b = st.columns(2)
    with col_a:
        connect_btn = st.button("🔗 Connect", use_container_width=True, type="primary")
    with col_b:
        disc_btn = st.button("✖ Disconnect", use_container_width=True)

    if connect_btn:
        st.session_state.host = host
        st.session_state.port = int(port)
        with st.spinner("Connecting to Tally..."):
            xml_resp = tally_post(build_company_list_xml(), host, int(port))
        if xml_resp is None:
            st.error("❌ Cannot reach Tally. Make sure Tally is open and the port is correct.")
            st.session_state.connected = False
        else:
            companies = parse_companies(xml_resp)
            st.session_state.companies = companies
            st.session_state.connected = True
            st.success(f"✅ Connected! Found {len(companies)} compan{'y' if len(companies)==1 else 'ies'}.")

    if disc_btn:
        st.session_state.connected = False
        st.session_state.companies = []
        st.info("Disconnected from Tally.")

    st.divider()
    if st.session_state.connected:
        st.markdown('<span class="badge-connected">● CONNECTED</span>', unsafe_allow_html=True)
        st.caption(f"`{host}:{port}`")
    else:
        st.markdown('<span class="badge-disconnected">● DISCONNECTED</span>', unsafe_allow_html=True)

    st.divider()
    st.markdown("### 📋 Instructions")
    st.markdown("""
1. Open **Tally Prime / ERP 9**
2. Enable **XML/HTTP Server** in Tally  
   *(Gateway → F12 → Advanced Config → Enable ODBC / HTTP)*
3. Note the **port** (default: **9000**)
4. Click **Connect** above
5. Select companies & reports below
""")


# ─────────────────────────────────────────────
#  MAIN  — HEADER
# ─────────────────────────────────────────────
st.markdown("""
<div style="display:flex; align-items:center; gap:14px; margin-bottom:24px;">
  <div style="width:48px;height:48px;border-radius:12px;
              background:linear-gradient(135deg,#4fffb0,#5b8fff);
              display:flex;align-items:center;justify-content:center;
              font-size:24px;">📊</div>
  <div>
    <h1 style="margin:0;font-size:26px;">Tally Data Extractor</h1>
    <p style="margin:0;color:#6b7280;font-size:13px;">
      Extract Ledgers · P&amp;L · Balance Sheet · Vouchers → Excel
    </p>
  </div>
</div>
""", unsafe_allow_html=True)

if not st.session_state.connected:
    st.markdown("""
    <div class="info-box">
      👈  Connect to Tally using the <strong>sidebar</strong> on the left to get started.
      Make sure <strong>Tally Prime</strong> or <strong>Tally ERP 9</strong> is running on this device with the XML/HTTP server enabled.
    </div>
    """, unsafe_allow_html=True)

    st.markdown("### How it works")
    c1, c2, c3, c4 = st.columns(4)
    for col, icon, step, desc in [
        (c1,"🔌","1. Connect","Enter your Tally host & port, then click Connect"),
        (c2,"🏢","2. Select Company","Choose the open company you want to extract data from"),
        (c3,"📑","3. Pick Reports","Select P&L, Balance Sheet, Ledgers, or Vouchers"),
        (c4,"📥","4. Export Excel","Download a formatted multi-sheet Excel workbook"),
    ]:
        with col:
            st.markdown(f"""
            <div class="tally-card" style="text-align:center;">
              <div style="font-size:32px;margin-bottom:8px;">{icon}</div>
              <div style="font-weight:700;font-size:14px;color:#e8eaf0;">{step}</div>
              <div style="font-size:12px;color:#6b7280;margin-top:6px;">{desc}</div>
            </div>""", unsafe_allow_html=True)
    st.stop()


# ─────────────────────────────────────────────
#  STEP 1 — COMPANY SELECTION
# ─────────────────────────────────────────────
st.markdown("## 🏢 Step 1 — Select Company")

companies = st.session_state.companies
if not companies:
    st.markdown('<div class="warn-box">⚠️ No companies found in Tally. Make sure at least one company is open in Tally.</div>', unsafe_allow_html=True)
    st.stop()

selected_company = st.selectbox(
    "Choose a company (only companies currently open in Tally are listed)",
    options=companies,
    help="Tally only exposes companies that are currently loaded/open."
)

# Metrics row
m1, m2, m3 = st.columns(3)
with m1:
    st.markdown(f"""<div class="metric-box">
      <div class="metric-value">{len(companies)}</div>
      <div class="metric-label">Open Companies</div>
    </div>""", unsafe_allow_html=True)
with m2:
    st.markdown(f"""<div class="metric-box">
      <div class="metric-value" style="color:#5b8fff;">{selected_company[:12]}{'…' if len(selected_company)>12 else ''}</div>
      <div class="metric-label">Selected</div>
    </div>""", unsafe_allow_html=True)
with m3:
    st.markdown(f"""<div class="metric-box">
      <div class="metric-value" style="color:#ffd166;">{st.session_state.port}</div>
      <div class="metric-label">Tally Port</div>
    </div>""", unsafe_allow_html=True)

st.divider()


# ─────────────────────────────────────────────
#  STEP 2 — DATE RANGE
# ─────────────────────────────────────────────
st.markdown("## 📅 Step 2 — Date Range")
d1, d2 = st.columns(2)
with d1:
    from_date = st.date_input("From Date", value=date(date.today().year, 4, 1))
with d2:
    to_date   = st.date_input("To Date",   value=date.today())

from_str = from_date.strftime("%Y%m%d")
to_str   = to_date.strftime("%Y%m%d")
st.caption(f"Tally format: `{from_str}` → `{to_str}`")

st.divider()


# ─────────────────────────────────────────────
#  STEP 3 — REPORT SELECTION
# ─────────────────────────────────────────────
st.markdown("## 📑 Step 3 — Choose Reports to Extract")

col_r1, col_r2, col_r3, col_r4 = st.columns(4)
with col_r1:
    do_ledger   = st.checkbox("📒 Ledgers",           value=True)
with col_r2:
    do_pl       = st.checkbox("📈 Profit & Loss",     value=True)
with col_r3:
    do_bs       = st.checkbox("⚖️ Balance Sheet",     value=True)
with col_r4:
    do_vouchers = st.checkbox("🧾 Voucher Register",  value=False)

st.divider()


# ─────────────────────────────────────────────
#  STEP 4 — LEDGER COLUMN CUSTOMIZER
# ─────────────────────────────────────────────
LEDGER_ALL_COLUMNS = [
    "Ledger Name", "Group / Parent", "Opening Balance",
    "Closing Balance", "GSTIN", "PAN", "Email", "Phone", "Address"
]

ledger_columns = LEDGER_ALL_COLUMNS  # default

if do_ledger:
    st.markdown("## 🗂️ Step 4 — Customise Ledger Columns")
    st.markdown('<div class="info-box">Select and reorder the columns you want in the Ledger sheet of the exported Excel file.</div>', unsafe_allow_html=True)
    st.markdown("")

    ledger_columns = st.multiselect(
        "Ledger columns to export (drag to reorder in the list below)",
        options=LEDGER_ALL_COLUMNS,
        default=LEDGER_ALL_COLUMNS,
        help="Uncheck columns you don't need. The order here is the order in Excel."
    )

    if not ledger_columns:
        st.warning("⚠️ No ledger columns selected — all columns will be exported.")
        ledger_columns = LEDGER_ALL_COLUMNS

    st.markdown("**Column order preview:**")
    cols_preview = st.columns(len(ledger_columns))
    for i, cname in enumerate(ledger_columns):
        cols_preview[i].markdown(f"""
        <div style="background:#1c2236;border:1px solid #252d3d;border-radius:6px;
                    padding:6px 10px;text-align:center;font-size:11px;color:#4fffb0;font-weight:600;">
          {i+1}. {cname}
        </div>""", unsafe_allow_html=True)

    st.divider()


# ─────────────────────────────────────────────
#  STEP 5 — EXTRACT & EXPORT
# ─────────────────────────────────────────────
st.markdown("## 🚀 Step 5 — Extract & Export")

if not (do_ledger or do_pl or do_bs or do_vouchers):
    st.markdown('<div class="warn-box">⚠️ Please select at least one report type above.</div>', unsafe_allow_html=True)
else:
    extract_btn = st.button("⬇️  Extract Data from Tally", type="primary", use_container_width=True)

    if extract_btn:
        sheets: dict = {}
        progress = st.progress(0, text="Starting extraction…")
        total_steps = sum([do_ledger, do_pl, do_bs, do_vouchers])
        step = 0

        # ── LEDGERS ──
        if do_ledger:
            progress.progress(int(step/total_steps*90), text="Fetching Ledgers…")
            xml = tally_post(build_ledger_xml(selected_company, from_str, to_str),
                             st.session_state.host, st.session_state.port)
            if xml:
                df_ledger = parse_ledgers(xml)
                sheets["Ledgers"] = (
                    f"Ledger Master — {selected_company}",
                    df_ledger,
                    ledger_columns
                )
            else:
                st.error("Failed to fetch Ledger data from Tally.")
            step += 1

        # ── P&L ──
        if do_pl:
            progress.progress(int(step/total_steps*90), text="Fetching Profit & Loss…")
            xml = tally_post(build_pl_xml(selected_company, from_str, to_str),
                             st.session_state.host, st.session_state.port)
            if xml:
                df_income, df_expense = parse_pl(xml)
                sheets["P&L Income"]   = (f"P&L — Income  [{from_date} to {to_date}]",  df_income,  None)
                sheets["P&L Expenses"] = (f"P&L — Expenses [{from_date} to {to_date}]", df_expense, None)
            else:
                st.error("Failed to fetch P&L data from Tally.")
            step += 1

        # ── BALANCE SHEET ──
        if do_bs:
            progress.progress(int(step/total_steps*90), text="Fetching Balance Sheet…")
            xml = tally_post(build_bs_xml(selected_company, to_str),
                             st.session_state.host, st.session_state.port)
            if xml:
                df_liab, df_assets = parse_bs(xml)
                sheets["BS Liabilities"] = (f"Balance Sheet — Liabilities (as on {to_date})", df_liab,   None)
                sheets["BS Assets"]      = (f"Balance Sheet — Assets (as on {to_date})",       df_assets, None)
            else:
                st.error("Failed to fetch Balance Sheet data from Tally.")
            step += 1

        # ── VOUCHERS ──
        if do_vouchers:
            progress.progress(int(step/total_steps*90), text="Fetching Voucher Register…")
            xml = tally_post(build_voucher_xml(selected_company, from_str, to_str),
                             st.session_state.host, st.session_state.port)
            if xml:
                df_vouchers = parse_vouchers(xml)
                sheets["Vouchers"] = (f"Voucher Register [{from_date} to {to_date}]", df_vouchers, None)
            else:
                st.error("Failed to fetch Voucher Register from Tally.")
            step += 1

        progress.progress(95, text="Building Excel workbook…")

        if sheets:
            excel_bytes = build_excel(sheets)
            progress.progress(100, text="✅ Done!")
            filename = f"Tally_{selected_company.replace(' ','_')}_{to_str}.xlsx"

            st.success(f"✅ Extracted **{len(sheets)} sheet(s)** successfully!")

            # Preview tables
            with st.expander("👁️ Preview extracted data", expanded=True):
                tab_names = list(sheets.keys())
                tabs = st.tabs(tab_names)
                for tab, sheet_name in zip(tabs, tab_names):
                    with tab:
                        _, df_preview, cols = sheets[sheet_name]
                        if cols:
                            df_preview = df_preview[[c for c in cols if c in df_preview.columns]]
                        st.dataframe(df_preview, use_container_width=True, height=300)
                        st.caption(f"{len(df_preview)} rows × {len(df_preview.columns)} columns")

            # Download button
            st.download_button(
                label="📥  Download Excel File",
                data=excel_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
        else:
            progress.empty()
            st.error("No data could be extracted. Check Tally connection and try again.")


# ─────────────────────────────────────────────
#  FOOTER
# ─────────────────────────────────────────────
st.divider()
st.markdown("""
<div style="text-align:center;color:#3a3f52;font-size:12px;padding:16px 0;">
  Tally Data Extractor &nbsp;|&nbsp; Communicates via Tally's built-in XML/HTTP server on port 9000
  &nbsp;|&nbsp; Supports Tally Prime &amp; Tally ERP 9
</div>
""", unsafe_allow_html=True)
